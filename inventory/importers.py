from __future__ import annotations

import csv
import datetime as dt
import hashlib
import ipaddress
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import load_workbook


def _normalize(value: object) -> str:
    """Return a trimmed string representation, defaulting to empty."""
    if value is None:
        return ""
    return str(value).strip()


def _compact_spaces(value: str) -> str:
    """Collapse multiple spaces to make comparisons easier."""
    return " ".join(value.split())


def _safe_ip(value: str) -> Optional[str]:
    """Validate an IP string; return None when invalid or empty."""
    value = _normalize(value)
    if not value:
        return None
    try:
        ipaddress.ip_address(value)
        return value
    except ValueError:
        return None


def _parse_bool(value: str) -> bool:
    return _normalize(value).lower() in {"true", "1", "si", "sí", "yes"}


def _parse_int(value: str) -> Optional[int]:
    value = _normalize(value)
    if not value:
        return None
    try:
        # Some exports use comma as thousands separator or decimal mark.
        cleaned = value.replace(".", "").replace(",", "")
        return int(cleaned)
    except ValueError:
        return None


def _parse_date_from_text(value: str) -> Optional[dt.date]:
    """
    Attempt to extract a date from free text (e.g. ' / /2025').
    Returns None when day/month are missing.
    """
    value = _normalize(value)
    if not value:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            parsed = dt.datetime.strptime(value, fmt).date()
            return parsed
        except ValueError:
            continue
    return None


@dataclass
class EquipmentItemData:
    section_label: str
    description: str
    brand_model: str
    serial_number: str
    units: str
    unit_status: str
    delivered: str
    observations: str
    raw_row: Dict[str, object]


@dataclass
class EquipmentRegisterData:
    source_name: str
    checksum: str
    service_date_text: str
    service_date: Optional[dt.date]
    service_order: str
    deployment: str
    allanamiento: str
    police_procedure: str
    other_notes: str
    raw_metadata: Dict[str, object]
    items: List[EquipmentItemData]


def parse_equipment_register(path: Path) -> EquipmentRegisterData:
    """
    Parse an equipment delivery/receipt spreadsheet (ANEXO VI).
    Extracts header metadata and item rows, keeping raw cells for traceability.
    """
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))

    def find_metadata_value(keys: Sequence[str]) -> str:
        for row in rows:
            for cell in row:
                if isinstance(cell, str):
                    for key in keys:
                        if key in cell:
                            _, _, tail = cell.partition(":")
                            return _compact_spaces(tail)
        return ""

    service_date_text = find_metadata_value(["FECHA DE SERVICIO"])
    service_order = find_metadata_value(["ORDEN DE SERVICIO"])
    deployment = find_metadata_value(["DESPLIEGUE"])
    allanamiento = find_metadata_value(["ALLANAMIENTO"])
    police_procedure = find_metadata_value(["PROCEDIMIENTO POLICIAL"])
    other_notes = find_metadata_value(["OTROS"])

    header_idx = None
    header_map: Dict[str, int] = {}
    for idx, row in enumerate(rows):
        normalized = [_normalize(cell).upper() for cell in row]
        if "DESCRIPCIÓN" in normalized and "MARCA/MODELO" in normalized:
            header_idx = idx
            header_map = {
                "section": 0,
                "description": normalized.index("DESCRIPCIÓN"),
                "brand": normalized.index("MARCA/MODELO"),
            }
            # Serial header varies; pick the first column that contains "SERIE" or "IDENTIFICACION"
            serial_positions = [i for i, value in enumerate(normalized) if "SERIE" in value or "IDENTIFICACION" in value]
            header_map["serial"] = serial_positions[0] if serial_positions else header_map["brand"] + 1
            unit_positions = [i for i, value in enumerate(normalized) if "UNIDADES" in value]
            header_map["units"] = unit_positions[0] if unit_positions else header_map["brand"] + 2
            status_positions = [i for i, value in enumerate(normalized) if "ESTADO" in value]
            header_map["status"] = status_positions[0] if status_positions else header_map["units"] + 1
            delivered_positions = [i for i, value in enumerate(normalized) if "ENTREGA" in value]
            header_map["delivered"] = delivered_positions[0] if delivered_positions else header_map["status"] + 1
            observations_positions = [i for i, value in enumerate(normalized) if "OBSERVACIONES" in value]
            header_map["observations"] = observations_positions[0] if observations_positions else header_map["delivered"] + 1
            break

    items: List[EquipmentItemData] = []
    if header_idx is not None:
        consecutive_blank_rows = 0
        for row in rows[header_idx + 1 :]:
            normalized = [_normalize(cell) for cell in row]
            if {"DESCRIPCIÓN", "MARCA/MODELO"} & set(value.upper() for value in normalized if value):
                # Skip repeated headers in the same sheet.
                continue

            relevant_values = [
                normalized[header_map["description"]] if len(normalized) > header_map["description"] else "",
                normalized[header_map["brand"]] if len(normalized) > header_map["brand"] else "",
                normalized[header_map["serial"]] if len(normalized) > header_map["serial"] else "",
            ]
            if not any(relevant_values):
                consecutive_blank_rows += 1
                if consecutive_blank_rows >= 3:
                    break
                continue
            consecutive_blank_rows = 0

            section = normalized[header_map["section"]] if len(normalized) > header_map["section"] else ""
            description = _compact_spaces(relevant_values[0])
            brand_model = _compact_spaces(relevant_values[1])
            serial_number = _compact_spaces(relevant_values[2])
            units = _compact_spaces(normalized[header_map["units"]]) if len(normalized) > header_map["units"] else ""
            unit_status = _compact_spaces(normalized[header_map["status"]]) if len(normalized) > header_map["status"] else ""
            delivered = _compact_spaces(normalized[header_map["delivered"]]) if len(normalized) > header_map["delivered"] else ""
            observations = _compact_spaces(normalized[header_map["observations"]]) if len(normalized) > header_map["observations"] else ""

            if not any([description, brand_model, serial_number, units, unit_status, delivered, observations]):
                continue

            items.append(
                EquipmentItemData(
                    section_label=_compact_spaces(section),
                    description=description,
                    brand_model=brand_model,
                    serial_number=serial_number,
                    units=units,
                    unit_status=unit_status,
                    delivered=delivered,
                    observations=observations,
                    raw_row={"row": normalized},
                )
            )

    checksum = hashlib.md5(path.read_bytes()).hexdigest()
    return EquipmentRegisterData(
        source_name=path.name,
        checksum=checksum,
        service_date_text=service_date_text,
        service_date=_parse_date_from_text(service_date_text),
        service_order=service_order,
        deployment=deployment,
        allanamiento=allanamiento,
        police_procedure=police_procedure,
        other_notes=other_notes,
        raw_metadata={"service_row": service_date_text, "order_row": service_order},
        items=items,
    )


@dataclass
class CameraInventoryRowData:
    source_name: str
    server_name: str
    device_name: str
    vendor: str
    model: str
    location: str
    logical_id: str
    device_id: str
    camera_id: str
    ip_address: Optional[str]
    mac_address: str
    firmware_version: str
    firmware_required: str
    serial_number: str
    connected: bool
    visible: bool
    error_indicators: str
    state: str
    bitrate_kbps: Optional[int]
    resolution: str
    quality: str
    frame_rate: str
    encryption: str
    retention: str
    appearance_search: str
    face_recognition: str
    raw_row: Dict[str, object]


def parse_camera_inventory_csv(path: Path) -> List[CameraInventoryRowData]:
    """
    Parse a camera inventory CSV exported from Avigilon.
    """
    field_map = {
        "Nombre de servidor": "server_name",
        "Nombre del dispositivo": "device_name",
        "Realizar": "vendor",
        "Modelo": "model",
        "Ubicacion": "location",
        "ID logico": "logical_id",
        "ID de dispositivo": "device_id",
        "Cadena de ID de la cámara": "camera_id",
        "Direccion IP": "ip_address",
        "Direccion MAC": "mac_address",
        "Version del firmware": "firmware_version",
        "Version del firmware obligatoria": "firmware_required",
        "Numero de serie": "serial_number",
        "Conectado": "connected",
        "Visible": "visible",
        "Indicadores de error": "error_indicators",
        "Estado": "state",
        "Velocidad de bits": "bitrate_kbps",
        "Resolucion": "resolution",
        "Calidad": "quality",
        "Velocidad de imagen": "frame_rate",
        "Cifrado": "encryption",
        "Retencion": "retention",
        "Appearance Search:": "appearance_search",
        "Reconocimiento facial:": "face_recognition",
    }

    with path.open("r", encoding="utf-8", errors="ignore") as fh:
        reader = csv.DictReader(fh, delimiter=";")
        rows: List[CameraInventoryRowData] = []
        for raw_row in reader:
            normalized_row = {_normalize(k): v for k, v in raw_row.items()}
            data: Dict[str, object] = {field_map.get(key, key): value for key, value in normalized_row.items()}
            ip_address = _safe_ip(data.get("ip_address", ""))  # type: ignore[arg-type]
            rows.append(
                CameraInventoryRowData(
                    source_name=path.name,
                    server_name=_normalize(data.get("server_name", "")),
                    device_name=_normalize(data.get("device_name", "")),
                    vendor=_normalize(data.get("vendor", "")),
                    model=_normalize(data.get("model", "")),
                    location=_normalize(data.get("location", "")),
                    logical_id=_normalize(data.get("logical_id", "")),
                    device_id=_normalize(data.get("device_id", "")),
                    camera_id=_normalize(data.get("camera_id", "")),
                    ip_address=ip_address,
                    mac_address=_normalize(data.get("mac_address", "")),
                    firmware_version=_normalize(data.get("firmware_version", "")),
                    firmware_required=_normalize(data.get("firmware_required", "")),
                    serial_number=_normalize(data.get("serial_number", "")),
                    connected=_parse_bool(data.get("connected", "")),  # type: ignore[arg-type]
                    visible=_parse_bool(data.get("visible", "")),  # type: ignore[arg-type]
                    error_indicators=_normalize(data.get("error_indicators", "")),
                    state=_normalize(data.get("state", "")),
                    bitrate_kbps=_parse_int(data.get("bitrate_kbps", "")),  # type: ignore[arg-type]
                    resolution=_normalize(data.get("resolution", "")),
                    quality=_normalize(data.get("quality", "")),
                    frame_rate=_normalize(data.get("frame_rate", "")),
                    encryption=_normalize(data.get("encryption", "")),
                    retention=_normalize(data.get("retention", "")),
                    appearance_search=_normalize(data.get("appearance_search", "")),
                    face_recognition=_normalize(data.get("face_recognition", "")),
                    raw_row=data,
                )
            )
    return rows
